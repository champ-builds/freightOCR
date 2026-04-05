import asyncio
import io
import json
import logging
import re
import tempfile
import traceback
import uuid as uuid_module
import zipfile
from pathlib import Path
from urllib.parse import unquote

import pandas as pd
import pdfplumber
from fastapi import FastAPI, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Base directory (so paths work on any deployment platform)
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent

app = FastAPI(title="Freight Bill Extractor")

# ---------------------------------------------------------------------------
# CORS — expose custom header to browser JS
# ---------------------------------------------------------------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Extraction-Summary"],
)

# ---------------------------------------------------------------------------
# ✅ Serve frontend
# ---------------------------------------------------------------------------
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

@app.get("/")
def serve_frontend():
    return FileResponse(str(BASE_DIR / "templates" / "index.html"))


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
COLUMN_NAMES = [
    "Bill No", "Bill Date",
    "Invoice Number", "Invoice Date", "Shipment Number", "Shipment Date",
    "Truck Number", "Gross Weight (TN)", "FI Document Number",
    "Freight (996791)", "Loading (996791)", "Unloading", "Multi Drop (996791)",
    "Handling", "Fixed PDP Remuneration", "Other Charges", "Total", "Exempted Items",
]

NUMERIC_COLS = [
    "Gross Weight (TN)", "Freight (996791)", "Loading (996791)", "Unloading",
    "Multi Drop (996791)", "Handling", "Fixed PDP Remuneration",
    "Other Charges", "Total", "Exempted Items",
]

DATE_COLS = ["Invoice Date", "Shipment Date"]


# ---------------------------------------------------------------------------
# PDF Extraction
# ---------------------------------------------------------------------------
def extract_freight_bill(pdf_path: str) -> pd.DataFrame:
    logger.info(f"Opening PDF: {pdf_path}")
    bill_no, bill_date = None, None
    all_rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""

            if page_num == 1:
                for line in text.splitlines():
                    if re.search(r"bill\s*no", line, re.IGNORECASE):
                        m = re.search(r":\s*(\d+)", line)
                        if m:
                            bill_no = m.group(1).strip()

                    if re.search(r"bill\s*date", line, re.IGNORECASE):
                        m = re.search(r":\s*(\d{2}-\d{2}-\d{4})", line)
                        if m:
                            bill_date = pd.to_datetime(
                                m.group(1), format="%d-%m-%Y", errors="coerce"
                            )

            tables = page.extract_tables() or []

            for table in tables:
                for row in table:
                    if not row:
                        continue
                    first = str(row[0]).strip() if row[0] else ""
                    if re.match(r"^\d{7,}", first):
                        padded = (row + [None] * 16)[:16]
                        all_rows.append([bill_no, bill_date] + padded)

    if not all_rows:
        raise ValueError("No valid invoice rows found in the PDF.")

    df = pd.DataFrame(all_rows, columns=COLUMN_NAMES)

    for col in DATE_COLS:
        df[col] = pd.to_datetime(df[col], format="%d-%m-%Y", errors="coerce")

    for col in NUMERIC_COLS:
        df[col] = pd.to_numeric(
            df[col].astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce",
        )

    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel Writing
# ---------------------------------------------------------------------------
def sanitize_sheet_name(name: str) -> str:
    invalid = r'[\\/*?:\[\]]'
    name = re.sub(invalid, "_", name)
    return name[:31]


def write_excel(results: dict[str, pd.DataFrame]) -> io.BytesIO:
    logger.info(f"Building Excel workbook with {len(results)} sheet(s)")
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F5496")

    for sheet_name, df in results.items():
        ws = wb.create_sheet(title=sanitize_sheet_name(sheet_name))

        # Headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if isinstance(value, pd.Timestamp):
                    cell.value = value.strftime("%d-%m-%Y") if not pd.isnull(value) else None
                else:
                    cell.value = None if str(value) in ("nan", "NaT", "None") else value

        # Column width
        for col_idx, col_name in enumerate(df.columns, start=1):
            lengths = [len(str(col_name))]
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                lengths.append(len(str(val or "")))
            max_len = max(lengths) if lengths else 10
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Core Processing
# ---------------------------------------------------------------------------
def process_pdf_files(pdf_files: list[Path]):
    successful, failed, extracted = [], [], {}

    for pdf_file in pdf_files:
        try:
            df = extract_freight_bill(str(pdf_file))
            extracted[pdf_file.stem] = df
            successful.append(pdf_file.stem)
        except Exception as exc:
            failed.append({
                "file": pdf_file.name,
                "reason": str(exc),
                "detail": traceback.format_exc().splitlines()[-1],
            })

    return extracted, successful, failed


def build_response(extracted, successful, failed):
    if not extracted:
        return JSONResponse(
            status_code=422,
            content={"message": "No PDFs extracted", "failed": failed},
        )

    excel_bytes = write_excel(extracted)

    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=freight_bills.xlsx",
            "X-Extraction-Summary": json.dumps({
                "successful": successful,
                "failed": failed
            }),
        },
    )


# ---------------------------------------------------------------------------
# Endpoint: Folder
# ---------------------------------------------------------------------------
@app.get("/extract")
def extract_pdfs(folder: str = Query(...)):
    try:
        folder = unquote(folder)
        logger.info(f"Request received for folder: {folder}")

        folder_path = Path(folder)

        if not folder_path.exists():
            return JSONResponse(status_code=400, content={"error": "Folder not found"})

        pdf_files = sorted(folder_path.glob("*.pdf"))

        if not pdf_files:
            return JSONResponse(status_code=404, content={"error": "No PDFs found"})

        extracted, successful, failed = process_pdf_files(pdf_files)
        return build_response(extracted, successful, failed)

    except Exception as e:
        logger.error(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Endpoint: ZIP Upload
# ---------------------------------------------------------------------------
@app.post("/extract-zip")
async def extract_zip(file: UploadFile = File(...)):
    try:
        if not file.filename.endswith(".zip"):
            return JSONResponse(status_code=400, content={"error": "Upload a ZIP file"})

        zip_bytes = await file.read()

        with tempfile.TemporaryDirectory() as tmpdir:
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                zf.extractall(tmpdir)

            pdf_files = list(Path(tmpdir).rglob("*.pdf"))

            if not pdf_files:
                return JSONResponse(status_code=404, content={"error": "No PDFs in ZIP"})

            extracted, successful, failed = process_pdf_files(pdf_files)
            return build_response(extracted, successful, failed)

    except Exception as e:
        logger.error(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Endpoint: ZIP Upload with SSE Progress Streaming
# ---------------------------------------------------------------------------
_excel_store: dict[str, bytes] = {}


@app.post("/extract-zip-stream")
async def extract_zip_stream(file: UploadFile = File(...)):
    if not file.filename.endswith(".zip"):
        return JSONResponse(status_code=400, content={"error": "Upload a ZIP file"})

    zip_bytes = await file.read()

    if not zipfile.is_zipfile(io.BytesIO(zip_bytes)):
        return JSONResponse(status_code=400, content={"error": "Invalid ZIP archive"})

    async def event_stream():
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                    zf.extractall(tmpdir)

                pdf_files = sorted(Path(tmpdir).rglob("*.pdf"))
                total = len(pdf_files)

                if total == 0:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'No PDFs found in ZIP'})}\n\n"
                    return

                successful, failed, extracted = [], [], {}

                for idx, pdf_file in enumerate(pdf_files, start=1):
                    # Send progress event BEFORE processing
                    yield f"data: {json.dumps({'type': 'progress', 'current': idx, 'total': total, 'file': pdf_file.name})}\n\n"
                    await asyncio.sleep(0.01)  # flush the event

                    try:
                        df = await asyncio.to_thread(
                            extract_freight_bill, str(pdf_file)
                        )
                        extracted[pdf_file.stem] = df
                        successful.append(pdf_file.stem)
                    except Exception as exc:
                        failed.append({
                            "file": pdf_file.name,
                            "reason": str(exc),
                        })

                if not extracted:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'No PDFs could be extracted', 'failed': failed})}\n\n"
                    return

                # Generate Excel and store for download
                excel_buf = write_excel(extracted)
                download_id = str(uuid_module.uuid4())
                _excel_store[download_id] = excel_buf.getvalue()

                yield f"data: {json.dumps({'type': 'complete', 'download_id': download_id, 'successful': successful, 'failed': failed})}\n\n"

        except Exception as e:
            logger.error(traceback.format_exc())
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ---------------------------------------------------------------------------
# Endpoint: Download Excel by ID
# ---------------------------------------------------------------------------
@app.get("/download/{download_id}")
def download_excel(download_id: str):
    if download_id not in _excel_store:
        return JSONResponse(status_code=404, content={"error": "Download not found or expired"})

    excel_bytes = _excel_store.pop(download_id)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=freight_bills.xlsx"},
    )